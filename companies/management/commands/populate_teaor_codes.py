# companies/management/commands/populate_teaor_codes.py
import openpyxl
from django.core.management.base import BaseCommand
from companies.models import TEORCode
from django.db import transaction

class Command(BaseCommand):
    help = 'Populates the TEORCode table from a given XLSX file (TEÁOR\'25 structure)'

    def add_arguments(self, parser):
        parser.add_argument('xlsx_file_path', type=str, help='The path to the TEÁOR XLSX file')
        parser.add_argument(
            '--sheet_name',
            type=str,
            help='Optional: Name of the sheet to process. If not provided, the active sheet is used.'
        )
        parser.add_argument(
            '--header_row',
            type=int,
            default=1, # Feltételezzük, hogy az első sor a fejléc, és az adatok a második sortól kezdődnek
            help='Row number of the header (1-indexed). Data is read from the next row.'
        )
        parser.add_argument(
            '--code_col',
            type=str,
            default='A', # Feltételezzük, hogy az 'A' oszlop a TEÁOR kód
            help='Column letter for TEÁOR codes (e.g., A, B).'
        )
        parser.add_argument(
            '--name_col',
            type=str,
            default='B', # Feltételezzük, hogy a 'B' oszlop a megnevezés
            help='Column letter for TEÁOR names (e.g., B, C).'
        )


    def handle(self, *args, **options):
        xlsx_file_path = options['xlsx_file_path']
        sheet_name = options['sheet_name']
        header_row_num = options['header_row']
        code_column_letter = options['code_col'].upper()
        name_column_letter = options['name_col'].upper()

        self.stdout.write(self.style.SUCCESS(f'Starting to populate TEOR codes from: {xlsx_file_path}'))

        created_count = 0
        updated_count = 0
        skipped_count = 0
        processed_rows = 0

        try:
            workbook = openpyxl.load_workbook(xlsx_file_path, read_only=True, data_only=True)
            
            if sheet_name:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                else:
                    self.stdout.write(self.style.ERROR(f"Sheet '{sheet_name}' not found in the workbook. Available sheets: {workbook.sheetnames}"))
                    return
            else:
                sheet = workbook.active # Az aktív (általában az első) munkalap
            
            self.stdout.write(self.style.SUCCESS(f"Processing sheet: '{sheet.title}'"))

            rows_iterator = sheet.iter_rows(min_row=header_row_num + 1) # Adatok olvasása a fejléc utáni sortól

            with transaction.atomic():
                for row_idx, row in enumerate(rows_iterator, start=header_row_num + 1):
                    processed_rows += 1
                    try:
                        # Oszlopbetűkből index konvertálás (pl. 'A' -> 0, 'B' -> 1)
                        # Az iter_rows cellákat ad vissza, így indexelhetünk.
                        # Vagy openpyxl.utils.column_index_from_string() használata
                        
                        # Értékek kiolvasása a cellákból a megadott oszlopbetűk alapján
                        # A cellák indexelése 0-alapú, de az oszlopbetűk 1-alapúak
                        # openpyxl.utils.cell.column_index_from_string() 1-alapú indexet ad
                        
                        code_val_cell = sheet[f"{code_column_letter}{row_idx}"]
                        name_val_cell = sheet[f"{name_column_letter}{row_idx}"]

                        code = str(code_val_cell.value).strip() if code_val_cell.value is not None else None
                        name = str(name_val_cell.value).strip() if name_val_cell.value is not None else None
                        
                        if code and name and 1 <= len(code) <= 10:
                            obj, created = TEORCode.objects.update_or_create(
                                code=code,
                                defaults={'name': name}
                            )
                            if created:
                                created_count += 1
                            else:
                                updated_count += 1
                        else:
                            if code or name: # Csak akkor logoljuk, ha volt valami adat, de nem volt elég
                                self.stdout.write(self.style.NOTICE(f"Skipping row {row_idx} due to missing/invalid code or name: Code='{code}', Name='{name}'"))
                            skipped_count += 1
                    except Exception as e:
                        self.stdout.write(self.style.ERROR(f"Error processing row {row_idx}: {row_values} - {e}"))
                        skipped_count += 1
                        continue
            
            workbook.close()

        except FileNotFoundError:
            self.stdout.write(self.style.ERROR(f"XLSX file not found at: {xlsx_file_path}"))
            return
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"An unexpected error occurred: {e}"))
            self.stdout.write(self.style.ERROR("Make sure 'openpyxl' is installed (pip install openpyxl) and the file is a valid XLSX."))
            return

        self.stdout.write(self.style.SUCCESS(f'Successfully processed {processed_rows} rows from the sheet.'))
        self.stdout.write(self.style.SUCCESS(f'Created {created_count} new TEOR codes.'))
        self.stdout.write(self.style.SUCCESS(f'Updated {updated_count} existing TEOR codes.'))
        if skipped_count > 0:
            self.stdout.write(self.style.WARNING(f'Skipped {skipped_count} rows (due to missing data or errors).'))