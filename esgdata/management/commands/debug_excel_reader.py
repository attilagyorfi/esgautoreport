# esgdata/management/commands/debug_excel_reader.py
import openpyxl
from django.core.management.base import BaseCommand, CommandError

class Command(BaseCommand):
    help = 'DEBUGGING SCRIPT: Reads and prints content from a questionnaire Excel sheet without saving to DB.'

    def add_arguments(self, parser):
        parser.add_argument('excel_file_path', type=str, help='The path to the questionnaire Excel file.')
        parser.add_argument('--sheet_name', type=str, required=True, help='Exact name of the sheet to process.')
        parser.add_argument('--header_row', type=int, default=1, help='Header row number (1-indexed). Data starts from the next row.')

    def handle(self, *args, **options):
        excel_file_path = options['excel_file_path']
        sheet_name = options['sheet_name']
        header_row_num = options['header_row']

        self.stdout.write(self.style.SUCCESS(f"--- EXCEL DEBUGGER INDÍTÁSA: {excel_file_path} ---"))
        self.stdout.write(f"Munkalap neve: '{sheet_name}'")

        try:
            workbook = openpyxl.load_workbook(excel_file_path, read_only=True)
            if sheet_name not in workbook.sheetnames:
                raise CommandError(f"A '{sheet_name}' munkalap nem található. Elérhető munkalapok: {workbook.sheetnames}")
            sheet = workbook[sheet_name]
            
            self.stdout.write(f"Munkalap sikeresen megnyitva: '{sheet.title}'. Maximális sor: {sheet.max_row}\n")

            rows_to_process = 0
            rows_with_data = 0
            
            # Iteráció a sorokon a fejléc után
            for row_idx, row_cells in enumerate(sheet.iter_rows(min_row=header_row_num + 1, max_row=sheet.max_row), start=header_row_num + 1):
                rows_to_process += 1
                
                # A, B, és D oszlopok olvasása (indexek: 0, 1, 3)
                sorszam = row_cells[0].value
                altem = row_cells[1].value
                kerdes = row_cells[3].value

                # A nyers cellatartalom kiírása
                self.stdout.write(f"Sor {row_idx}:  A oszlop='{sorszam}'   B oszlop='{altem}'   D oszlop='{kerdes}'")

                # Ellenőrzés, hogy az importáló szkriptünk validnak tekintené-e ezt a sort
                if sorszam and altem and kerdes:
                    rows_with_data += 1
                    self.stdout.write(self.style.SUCCESS(f"  -> Ez a sor érvényes kérdésnek tűnik."))
                else:
                    self.stdout.write(self.style.WARNING(f"  -> Ezt a sort az importáló szkript KIHAGYNÁ (hiányzó adat az A, B, vagy D oszlopban)."))

            self.stdout.write("\n" + self.style.SUCCESS("--- DEBUGGER BEFEJEZVE ---"))
            self.stdout.write(f"Beolvasott sorok száma (fejléc után): {rows_to_process}")
            self.stdout.write(self.style.SUCCESS(f"Érvényes kérdésként azonosított sorok: {rows_with_data}"))
            
            if rows_with_data == 0 and rows_to_process > 0:
                 self.stdout.write(self.style.ERROR("HIBA: Egyetlen érvényes kérdést sem talált a szkript! Kérjük, ellenőrizd az Excel fájl szerkezetét, az oszlopok tartalmát és a --header_row paramétert."))

        except FileNotFoundError:
            raise CommandError(f"A fájl nem található: {excel_file_path}")
        except Exception as e:
            raise CommandError(f"Hiba történt a fájl olvasása közben: {e}")