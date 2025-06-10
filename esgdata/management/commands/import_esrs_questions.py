# esgdata/management/commands/import_esrs_questions.py
import openpyxl  # Módosítva, hogy Excel fájlt olvasson
from django.core.management.base import BaseCommand, CommandError
from esgdata.models import ESGDataPoint
from django.db import transaction

class Command(BaseCommand):
    help = 'Imports questions from the structured ESRS questionnaire Excel file, skipping MNB-specific questions.'

    def add_arguments(self, parser):
        parser.add_argument('excel_file_path', type=str, help='The path to the questionnaire Excel file.')
        # JAVÍTÁS: Hozzáadjuk a hiányzó --sheet_name argumentumot
        parser.add_argument(
            '--sheet_name',
            type=str,
            required=True, # Legyen kötelező, mert tudnunk kell, melyik munkalapot olvassuk
            help='The exact name of the sheet to process inside the Excel file.'
        )
        parser.add_argument(
            '--clear',
            action='store_true',
            help='Deletes all existing ESGDataPoint entries before importing.',
        )

    def handle(self, *args, **options):
        excel_file_path = options['excel_file_path']
        sheet_name = options['sheet_name'] # Most már helyesen be tudja olvasni
        
        # Pillar és Response Type megfeleltetések
        pillar_mapping = {'E': 'environmental', 'S': 'social', 'G': 'governance'}
        response_type_mapping = {
            'legördülő lista (választás)': ESGDataPoint.DATATYPE_DROPDOWN,
            'igen/nem (logikai)': ESGDataPoint.DATATYPE_BOOLEAN,
            'szám': ESGDataPoint.DATATYPE_NUMBER,
            'szöveges kifejtés': ESGDataPoint.DATATYPE_TEXT,
            'dátum': ESGDataPoint.DATATYPE_DATE,
            'fájl feltöltés': ESGDataPoint.DATATYPE_FILE,
        }

        try:
            with transaction.atomic():
                if options['clear']:
                    self.stdout.write(self.style.WARNING('Deleting all existing ESGDataPoint objects...'))
                    deleted_count, _ = ESGDataPoint.objects.all().delete()
                    self.stdout.write(self.style.SUCCESS(f'Successfully deleted {deleted_count} ESGDataPoint objects.'))

                self.stdout.write(self.style.SUCCESS(f"Importing questions from '{excel_file_path}', sheet: '{sheet_name}'..."))
                
                workbook = openpyxl.load_workbook(excel_file_path, read_only=True, data_only=True)
                if sheet_name not in workbook.sheetnames:
                    raise CommandError(f"A '{sheet_name}' munkalap nem található. Elérhető munkalapok: {workbook.sheetnames}")
                sheet = workbook[sheet_name]

                # Fejléc beolvasása, hogy név alapján hivatkozhassunk az oszlopokra
                header = [cell.value for cell in sheet[1]]

                created_count = 0
                updated_count = 0
                skipped_mnb = 0
                skipped_other = 0

                for row_idx, row_cells in enumerate(sheet.iter_rows(min_row=2), start=2): # Feltételezve, hogy az adatok a 2. sortól kezdődnek
                    row_data = {header[i]: str(cell.value).strip() if cell.value is not None else None for i, cell in enumerate(row_cells)}

                    if row_data.get('ESG beszámoló egyezőség') == 'MNB kérdés':
                        skipped_mnb += 1
                        continue

                    q_identifier = row_data.get('Adatpont')
                    q_text = row_data.get('Kérdés')
                    pillar_char = row_data.get('Pillér')

                    if not q_identifier or not q_text or not pillar_char:
                        skipped_other += 1
                        continue
                    
                    pillar_key = pillar_mapping.get(pillar_char.upper())
                    if not pillar_key:
                        self.stdout.write(self.style.WARNING(f"Skipping row with identifier '{q_identifier}': Invalid pillar character '{pillar_char}'."))
                        skipped_other += 1
                        continue
                    
                    resp_type_str = (row_data.get('Válasz típusa') or 'szöveges kifejtés').lower().strip()
                    response_data_type = response_type_mapping.get(resp_type_str, ESGDataPoint.DATATYPE_TEXT)

                    datapoint, created = ESGDataPoint.objects.update_or_create(
                        question_number=q_identifier,
                        defaults={
                            'question_text': q_text,
                            'pillar': pillar_key,
                            'response_data_type': response_data_type,
                            'guidance': row_data.get('Útmutató', ''),
                            'choice_option_group': row_data.get('Legördülő lista azonosító') if response_data_type == ESGDataPoint.DATATYPE_DROPDOWN else None,
                            'is_voluntary': True if (row_data.get('Opcionális') or '').lower() == 'igen' else False,
                        }
                    )
                    if created: created_count += 1
                    else: updated_count += 1

                self.stdout.write(self.style.SUCCESS(f"\nImport finished!"))
                self.stdout.write(f"  Created: {created_count}")
                self.stdout.write(f"  Updated: {updated_count}")
                self.stdout.write(self.style.NOTICE(f"  Skipped (MNB question): {skipped_mnb}"))
                self.stdout.write(self.style.WARNING(f"  Skipped (Other reasons): {skipped_other}"))

        except FileNotFoundError:
            raise CommandError(f"File not found at: {excel_file_path}")
        except Exception as e:
            raise CommandError(f"An error occurred: {e}")