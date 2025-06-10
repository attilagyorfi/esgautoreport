# esgdata/management/commands/import_questionnaire_excel.py
import openpyxl
from django.core.management.base import BaseCommand, CommandError
from esgdata.models import ESGDataPoint, ChoiceOption
from django.db import transaction

class Command(BaseCommand):
    help = 'Imports E, S, G questions from a questionnaire Excel sheet, handling complex row structures.'

    def add_arguments(self, parser):
        parser.add_argument('excel_file_path', type=str, help='The path to the questionnaire Excel file.')
        parser.add_argument('--sheet_name', type=str, required=True, help='Name of the sheet to process.')
        parser.add_argument('--header_row', type=int, default=1, help='Header row number (1-indexed).')

    def get_cell_value(self, cell):
        return str(cell.value).strip() if cell.value is not None else None

    def handle(self, *args, **options):
        excel_file_path = options['excel_file_path']
        sheet_name = options['sheet_name']
        header_row_num = options['header_row']

        self.stdout.write(self.style.SUCCESS(f"Importing E,S,G questions from '{excel_file_path}', sheet: '{sheet_name}'..."))

        pillar_mapping = {'E': 'environmental', 'S': 'social', 'G': 'governance'}
        response_type_mapping = {
            'szöveges kifejtés': ESGDataPoint.DATATYPE_TEXT, 'szöveg': ESGDataPoint.DATATYPE_TEXT,
            'szám': ESGDataPoint.DATATYPE_NUMBER, 'igen/nem': ESGDataPoint.DATATYPE_BOOLEAN
        }
        
        created_count = 0
        updated_count = 0
        skipped_count = 0

        try:
            workbook = openpyxl.load_workbook(excel_file_path, read_only=True, data_only=True)
            if sheet_name not in workbook.sheetnames:
                raise CommandError(f"A '{sheet_name}' munkalap nem található. Elérhető munkalapok: {workbook.sheetnames}")
            sheet = workbook[sheet_name]
            
            with transaction.atomic():
                last_q_identifier = None  # Emlékszik az utolsó azonosítóra a B oszlopból

                for row_cells in sheet.iter_rows(min_row=header_row_num + 1):
                    # Mindig próbáljuk olvasni a B oszlopot
                    current_q_identifier = self.get_cell_value(row_cells[1])
                    if current_q_identifier:
                        last_q_identifier = current_q_identifier # Ha van új, frissítjük a memóriát

                    q_text = self.get_cell_value(row_cells[3]) # Kérdés szövege a D oszlopból

                    # Csak akkor dolgozunk, ha van kérdésszöveg ÉS van "megjegyzett" azonosítónk
                    if not q_text or not last_q_identifier:
                        skipped_count += 1
                        continue

                    pillar_char = last_q_identifier[0].upper()
                    pillar_key = pillar_mapping.get(pillar_char)

                    if not pillar_key:
                        self.stdout.write(self.style.WARNING(f"Skipping row with identifier '{last_q_identifier}': Invalid pillar character '{pillar_char}'."))
                        skipped_count += 1
                        continue
                    
                    # A többi adat kiolvasása és feldolgozása...
                    q_requirement = self.get_cell_value(row_cells[2]) or ""
                    q_answer_options_str = self.get_cell_value(row_cells[4]) or ""
                    q_extra_info = self.get_cell_value(row_cells[5]) or ""
                    
                    response_data_type = ESGDataPoint.DATATYPE_TEXT
                    choice_group = None
                    # ... (válasz típusának logikája változatlan) ...
                    if "igen/nem" in q_answer_options_str.lower():
                        response_data_type = ESGDataPoint.DATATYPE_BOOLEAN
                    elif "szám" in q_answer_options_str.lower():
                        response_data_type = ESGDataPoint.DATATYPE_NUMBER
                    elif q_answer_options_str and ("Igen," in q_answer_options_str or "Válasz:" in q_answer_options_str or "\n" in q_answer_options_str):
                        response_data_type = ESGDataPoint.DATATYPE_DROPDOWN
                        choice_group = f"CHOICE_GROUP_{last_q_identifier.replace('.', '_')}"


                    guidance_text = ""
                    if q_requirement: guidance_text += f"<p><strong>Követelmény:</strong> {q_requirement}</p>"
                    if q_extra_info: guidance_text += f"<p><strong>További információ:</strong> {q_extra_info}</p>"

                    # A question_number most már a megjegyzett azonosító lesz
                    datapoint, created = ESGDataPoint.objects.update_or_create(
                        question_number=last_q_identifier,
                        defaults={
                            'question_text': q_text,
                            'pillar': pillar_key,
                            'response_data_type': response_data_type,
                            'guidance': guidance_text,
                            'choice_option_group': choice_group,
                        }
                    )
                    
                    if response_data_type == ESGDataPoint.DATATYPE_DROPDOWN and choice_group:
                        options = [opt.strip() for opt in q_answer_options_str.replace("Válasz:", "").split('\n') if opt.strip()]
                        for option_text in options:
                            ChoiceOption.objects.get_or_create(group_identifier=choice_group, text=option_text)
                    
                    if created: created_count += 1
                    else: updated_count += 1
                
                self.stdout.write(self.style.SUCCESS(f"Import finished. Created: {created_count}, Updated: {updated_count}, Skipped: {skipped_count}"))

        except Exception as e:
            raise CommandError(f"An error occurred: {e}")