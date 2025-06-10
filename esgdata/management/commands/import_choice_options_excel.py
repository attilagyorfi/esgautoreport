import openpyxl
from django.core.management.base import BaseCommand, CommandError
from esgdata.models import ChoiceOption
from django.db import transaction

class Command(BaseCommand):
    help = 'Imports choice options from a specified Excel sheet into ChoiceOption model.'

    def add_arguments(self, parser):
        parser.add_argument('excel_file_path', type=str, help='The path to the choice options Excel file.')
        parser.add_argument('--sheet_name', type=str, help='Optional: Name of the sheet. Uses active sheet if not provided.')
        parser.add_argument('--header_row', type=int, default=1, help='Header row number (1-indexed).')
        parser.add_argument('--group_col', type=str, default='A', help='Column for Group Identifier.')
        parser.add_argument('--text_col', type=str, default='B', help='Column for Option Text.')

    def get_cell_value(self, sheet, row_num, col_letter):
        if col_letter:
            cell = sheet[f"{col_letter.upper()}{row_num}"]
            return str(cell.value).strip() if cell.value is not None else None
        return None

    def handle(self, *args, **options):
        excel_file_path = options['excel_file_path']
        sheet_name = options['sheet_name']
        header_row_num = options['header_row']
        group_col = options['group_col']
        text_col = options['text_col']

        self.stdout.write(self.style.SUCCESS(f"Importing choice options from '{excel_file_path}'..."))

        created_count = 0
        updated_count = 0
        skipped_count = 0
        processed_rows = 0

        try:
            workbook = openpyxl.load_workbook(excel_file_path, read_only=True, data_only=True)
            sheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
            self.stdout.write(f"Processing sheet: '{sheet.title}'")

            rows_iterator = sheet.iter_rows(min_row=header_row_num + 1)

            with transaction.atomic():
                for row_idx, row_cells in enumerate(rows_iterator, start=header_row_num + 1):
                    processed_rows += 1
                    group_identifier = self.get_cell_value(sheet, row_idx, group_col)
                    option_text = self.get_cell_value(sheet, row_idx, text_col)

                    if not group_identifier or not option_text:
                        self.stdout.write(self.style.WARNING(f"Skipping row {row_idx}: Missing group identifier or option text."))
                        skipped_count += 1
                        continue

                    obj, created = ChoiceOption.objects.update_or_create(
                        group_identifier=group_identifier,
                        text=option_text,
                        defaults={}
                    )
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

            workbook.close()
            self.stdout.write(self.style.SUCCESS(f"Import completed."))
            self.stdout.write(f"Processed {processed_rows} rows. Created: {created_count}, Updated: {updated_count}, Skipped: {skipped_count}")

        except FileNotFoundError:
            raise CommandError(f"Excel file not found at: {excel_file_path}")
        except Exception as e:
            raise CommandError(f"An error occurred: {e}")